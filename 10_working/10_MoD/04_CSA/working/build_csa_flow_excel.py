from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

base_dir = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\04_CSA")
out_path = base_dir / "CSA_Webex_フロー整理_20260526_01.xlsx"

wb = Workbook()

header_fill = PatternFill("solid", fgColor="006AA7")
header_font = Font(color="FFFFFF", bold=True)
wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def format_sheet(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(widths)):
        for c in row:
            c.alignment = wrap
            c.border = border
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font


# Sheet 1: スライド別フロー
ws1 = wb.active
ws1.title = "スライド別フロー"
ws1.append([
    "Slide", "スライドタイトル", "フローID", "利用者/端末", "通信区分", "想定経路", "CSA経由", "要点"
])

slide_flows = [
    (1, "Cisco Secure Access 基礎知識ガイド", "S1-F1", "全ユーザ", "全体像", "利用者→Webex/社内アプリ", "条件付き", "SSE全体像の導入スライド"),
    (2, "Agenda", "S2-F1", "全ユーザ", "説明順序", "要件整理→設計→運用", "条件付き", "以降の判定フローを示す目次"),
    (3, "CSAとは", "S3-F1", "Webex App / IPPhone", "アクセス制御", "端末→CSA→アプリ", "Yes", "ゼロトラストで統合保護"),
    (4, "6つの機能", "S4-F1", "Webex App", "Signaling", "Webex App→CSA(SWG/FWaaS)→Webex Cloud", "Yes", "SignalingはCSA適用しやすい"),
    (4, "6つの機能", "S4-F2", "Webex App", "Media", "Webex App→Internet直通→Webex Media", "No(推奨)", "品質優先で直通が基本"),
    (5, "音声/通話品質", "S5-F1", "IPPhone", "音声RTP/SRTP", "IPPhone→LGW/SBC→PSTN or Webex", "No(推奨)", "遅延・ジッタ抑制のため"),
    (5, "音声/通話品質", "S5-F2", "Webex App", "Signaling", "Webex App→CSA→Webex Control Plane", "Yes", "制御系はCSA経由可"),
    (6, "Webex利用時の役割", "S6-F1", "Webex App", "Signaling(HTTPS)", "端末→CSA(SWG/FWaaS/DNS)→Webex", "Yes", "可視化/制御の中心"),
    (6, "Webex利用時の役割", "S6-F2", "Webex App", "Media(RTP/SRTP)", "端末→Webex Media Node(直通)", "No(推奨)", "全面経由はPoCで検証"),
    (7, "ライセンス/課金", "S7-F1", "運用者", "契約", "SIA/SPAユーザ単位で設計", "N/A", "技術フローではなく費用フロー"),
    (8, "事例/SIPS・SRTP", "S8-F1", "LGW", "SIP/SRTP", "LGW→SIP Trunk/Cloud", "通常No", "通話品質優先。CSA全量経由事例は未確認"),
    (8, "事例/SIPS・SRTP", "S8-F2", "Webex App", "段階導入", "SignalingのみCSA→Mediaは直通", "Mixed", "段階導入を推奨"),
    (9, "設計ベストプラクティス", "S9-F1", "全端末", "設計判定", "Signaling=CSA / Media=直通", "Mixed", "KPIで継続評価"),
    (10, "まとめ", "S10-F1", "全体", "最終方針", "セキュリティと品質の分離最適化", "Mixed", "本資料の結論"),
]
for r in slide_flows:
    ws1.append(list(r))

format_sheet(ws1, [8, 28, 12, 16, 14, 36, 12, 36])

# Sheet 2: 端末別経路判定
ws2 = wb.create_sheet("端末別経路判定")
ws2.append([
    "端末", "ユーザシナリオ", "通信", "送信元", "宛先", "推奨経路", "CSA経由", "理由", "備考"
])

device_matrix = [
    ("Webex App", "サインイン/制御", "HTTPS/TLS", "ユーザPC", "Webex Control Plane", "Webex App→CSA→Webex", "Yes", "SWG/FWaaS/DNSで制御しやすい", "DEMで可視化"),
    ("Webex App", "通話メディア", "RTP/SRTP", "ユーザPC", "Webex Media", "Webex App→Internet直通", "No(推奨)", "遅延/ジッタ最小化", "全面経由はPoC判断"),
    ("IPPhone", "登録/制御", "SIP/TLS", "IPPhone", "CUCM/Webex Calling", "LAN/WAN既存経路", "通常No", "端末は既存音声設計を優先", "必要時のみ境界制御"),
    ("IPPhone", "通話メディア", "RTP/SRTP", "IPPhone", "対向端末/PSTN", "IPPhone→LGW/SBC→対向", "No(推奨)", "音声品質優先", "ローカルブレイクアウト推奨"),
    ("LGW", "PSTN接続", "SIP/SRTP", "LGW", "PSTN/SIP Trunk", "LGW→SIP Trunk", "No(推奨)", "音声GW区間は低遅延重視", "FW最小許可"),
    ("Webex App", "社内アプリ利用", "HTTPS/TCP", "ユーザPC", "Private App", "Webex App→CSA(ZTNA/VPNaaS)→Private App", "Yes", "ゼロトラスト適用", "最小権限アクセス"),
    ("管理者", "運用監視", "Telemetry", "端末/ネットワーク", "DEM", "各経路→DEM", "N/A", "品質・障害可視化", "KPI監視"),
]
for r in device_matrix:
    ws2.append(list(r))

format_sheet(ws2, [14, 16, 14, 14, 18, 34, 12, 28, 20])

# Sheet 3: 判定ルール
ws3 = wb.create_sheet("判定ルール")
ws3.append(["ルールID", "判定条件", "推奨", "CSA経由", "補足"])
rules = [
    ("R-01", "Webex Signaling (HTTPS/TLS)", "SWG/FWaaS/DNS経由で制御", "Yes", "URL許可・脅威遮断・可視化を優先"),
    ("R-02", "Webex Media (RTP/SRTP)", "直通経路を優先", "No(推奨)", "遅延・ジッタ・ロス低減"),
    ("R-03", "IPPhone 音声", "既存音声網/LGW経路を優先", "No(推奨)", "通話品質を最優先"),
    ("R-04", "LGW-PSTN 区間", "SIP Trunk直結", "No(推奨)", "音声GW処理を単純化"),
    ("R-05", "社内Private Appアクセス", "ZTNA/VPNaaS適用", "Yes", "最小権限アクセス"),
    ("R-06", "セキュリティ要件でMedia検査が必須", "PoCでKPI確認後に限定適用", "条件付き", "MOS/遅延/ジッタ閾値を満たすこと"),
]
for r in rules:
    ws3.append(list(r))
format_sheet(ws3, [10, 32, 28, 12, 36])

# Sheet 4: メタ情報
ws4 = wb.create_sheet("作成情報")
ws4.append(["項目", "値"])
ws4.append(["作成日", date.today().isoformat()])
ws4.append(["基準資料", "CSA_基礎知識.pptx (Cisco light 5 12 2025 テンプレート)" ] )
ws4.append(["用途", "端末種別ごとの CSA 経由/非経由フロー整理" ])
ws4.append(["対象端末", "Webex App / IPPhone / LGW" ])
ws4.append(["注意", "Media は品質優先で直通推奨。全量CSA経由はPoCで検証" ])
format_sheet(ws4, [18, 80])

wb.save(out_path)
print(f"保存完了: {out_path}")
