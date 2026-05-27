from openpyxl import load_workbook

path = r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\MoD_R08_PJ_AI-List.xlsx"
wb = load_workbook(path)

print("Sheets:", wb.sheetnames)

for ws in wb.worksheets:
    cfr = ws.conditional_formatting
    if not cfr:
        continue

    print(f"\n=== {ws.title} ===")
    if not cfr._cf_rules:
        print("(No rules)")
        continue

    for sqref, rules in cfr._cf_rules.items():
        for i, rule in enumerate(rules, start=1):
            formula = getattr(rule, "formula", None)
            rule_type = getattr(rule, "type", None)
            has_dxf = bool(getattr(rule, "dxf", None))
            print(f"Range={sqref} | Rule#{i} type={rule_type} formula={formula} dxf={has_dxf}")
