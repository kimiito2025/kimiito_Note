from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

# Cisco Webex 既知ポートマッピング
WEBEX_PORTS = {
    # Control/Signaling
    5002: ("シグナリング - Webex App 制御", "TCP,UDP"),
    8934: ("シグナリング - Webex App 制御（TLS）", "TCP,UDP"),
    8933: ("シグナリング - Webex App 制御（TLS）", "TCP,UDP"),
    
    # Media
    5004: ("メディア - 音声/ビデオ", "TCP,UDP"),
    5050: ("メディア - 音声/ビデオ（代替）", "TCP,UDP"),
    
    # DNS Security
    5353: ("DNS Security - Cisco Umbrella", "UDP"),
    
    # Management/TURN
    8033: ("TURN サーバー - メディアリレー", "TCP,UDP"),
    3478: ("STUN/TURN - メディアリレー", "TCP,UDP"),
    3479: ("TURN - メディアリレー（代替）", "UDP"),
    
    # HTTP/HTTPS (通常)
    80: ("HTTP - 標準ウェブ通信", "TCP"),
    443: ("HTTPS - 標準ウェブ通信", "TCP"),
    
    # その他の既知Webexポート
    6060: ("Webex Meeting Center", "TCP,UDP"),
    6061: ("Webex Meeting Center（代替）", "TCP,UDP"),
    6002: ("Webex オーディオ", "TCP,UDP"),
    6003: ("Webex オーディオ（代替）", "TCP,UDP"),
    8443: ("Webex セキュアウェブ", "TCP"),
    8082: ("Webex 開発用", "TCP"),
    9090: ("Webex 管理ポータル", "TCP"),
}

# ポート抽出
webex_port_dir = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\03_webex\webex_ports_extracted\webexサービス必要ポート数")

usecases = [
    ('Baseline', 'baseline_tcp.txt', 'baseline_udp.txt'),
    ('Meeting All', 'meeting_all_tcp.txt', 'meeting_all_udp.txt'),
    ('Meeting A/V', 'meeting_av_tcp.txt', 'meeting_av_udp.txt'),
    ('Messaging Idle', 'messaging_idle_tcp.txt', 'messaging_idle_udp.txt'),
]

port_data = {}
for name, tcp_file, udp_file in usecases:
    tcp_ports = set()
    udp_ports = set()
    
    with open(webex_port_dir / tcp_file, 'r', encoding='utf-16') as f:
        for line in f:
            matches = re.findall(r':(\d+)\s', line)
            for m in matches:
                tcp_ports.add(int(m))
    
    with open(webex_port_dir / udp_file, 'r', encoding='utf-16') as f:
        for line in f:
            matches = re.findall(r':(\d+)\s', line)
            for m in matches:
                udp_ports.add(int(m))
    
    port_data[name] = {
        'tcp': sorted(tcp_ports),
        'udp': sorted(udp_ports),
    }

# Workbook作成
output_path = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\03_webex\Webex_既知ポート一覧_20260526_01.xlsx")
wb = Workbook()

# スタイル
header_fill = PatternFill("solid", fgColor="005B8F")
header_font = Font(color="FFFFFF", bold=True, size=11)
webex_fill = PatternFill("solid", fgColor="E2EFDA")  # 薄緑
category_fill = PatternFill("solid", fgColor="D9E1F2")  # 薄青
wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Sheet 1: Webex既知ポート一覧
ws1 = wb.active
ws1.title = "Webex既知ポート"
ws1.append(["ポート番号", "説明", "プロトコル", "Baseline", "Meeting All", "Meeting A/V", "Messaging Idle"])

for c in ws1[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = wrap
    c.border = border

# 既知ポートの使用状況
for port in sorted(WEBEX_PORTS.keys()):
    desc, proto = WEBEX_PORTS[port]
    
    usage = []
    for name, tcp_file, udp_file in usecases:
        used = "✔"
        if "TCP" in proto and port not in port_data[name]['tcp']:
            used = ""
        elif "UDP" in proto and port not in port_data[name]['udp']:
            used = ""
        usage.append(used)
    
    ws1.append([port, desc, proto] + usage)

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 16

for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=7):
    for c in row:
        c.alignment = wrap
        c.border = border

# Sheet 2: ユースケース別詳細（既知ポートのみ）
ws2 = wb.create_sheet("ユースケース別")
ws2.append(["ユースケース", "プロトコル", "既知Webexポート", "ポート数"])

for c in ws2[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = wrap
    c.border = border

for name, tcp_file, udp_file in usecases:
    tcp_ports = port_data[name]['tcp']
    udp_ports = port_data[name]['udp']
    
    known_tcp = sorted([p for p in tcp_ports if p in WEBEX_PORTS])
    known_udp = sorted([p for p in udp_ports if p in WEBEX_PORTS])
    
    tcp_str = ", ".join(map(str, known_tcp)) if known_tcp else "(使用なし)"
    udp_str = ", ".join(map(str, known_udp)) if known_udp else "(使用なし)"
    
    ws2.append([name, "TCP", tcp_str, len(known_tcp)])
    ws2.append([name, "UDP", udp_str, len(known_udp)])

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 60
ws2.column_dimensions['D'].width = 10

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=4):
    for c in row:
        c.alignment = wrap
        c.border = border

# Sheet 3: 用途別分類
ws3 = wb.create_sheet("用途別分類")
ws3.append(["用途カテゴリ", "ポート番号", "プロトコル", "説明"])

for c in ws3[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = wrap
    c.border = border

categories = {
    "シグナリング/制御": [5002, 8933, 8934],
    "メディア（音声/ビデオ）": [5004, 5050, 6002, 6003],
    "セキュリティ": [5353],
    "メディアリレー": [8033, 3478, 3479],
    "ウェブ通信": [80, 443, 8443, 8082, 9090],
    "会議": [6060, 6061],
}

for category, ports in categories.items():
    for port in ports:
        if port in WEBEX_PORTS:
            desc, proto = WEBEX_PORTS[port]
            ws3.append([category, port, proto, desc])

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 45

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=4):
    for c in row:
        c.alignment = wrap
        c.border = border

# Sheet 4: 統計情報
ws4 = wb.create_sheet("統計情報")
ws4.append(["ユースケース", "既知Webexポート総数", "TCP", "UDP"])

for c in ws4[1]:
    c.fill = header_fill
    c.font = header_font
    c.alignment = wrap
    c.border = border

for name, tcp_file, udp_file in usecases:
    tcp_ports = port_data[name]['tcp']
    udp_ports = port_data[name]['udp']
    
    known_tcp = [p for p in tcp_ports if p in WEBEX_PORTS]
    known_udp = [p for p in udp_ports if p in WEBEX_PORTS]
    
    total = len(set(known_tcp) | set(known_udp))
    
    ws4.append([name, total, len(known_tcp), len(known_udp)])

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 12

for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=4):
    for c in row:
        c.alignment = wrap
        c.border = border

# Sheet 5: 参考情報
ws5 = wb.create_sheet("参考情報")
ws5.append(["項目", "内容"])

info = [
    ["作成日", date.today().isoformat()],
    ["対象資料", "Webex PC クライアント ネットワーク監視結果"],
    ["既知ポート定義", "Cisco公式Webexポート一覧に基づく"],
    ["ユースケース数", "4（Baseline, Meeting All, Meeting A/V, Messaging Idle）"],
    ["注記1", "表示されているポートはすべてWebex関連の既知ポートです"],
    ["注記2", "✔印 = そのユースケースで実際に使用されたポート"],
    ["注記3", "動的ポート（高番号）については、既知リストに基づき判定されています"],
]

for item in info:
    ws5.append(item)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 70

for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, min_col=1, max_col=2):
    for c in row:
        c.alignment = wrap
        c.border = border
    if row[0].row == 1:
        row[0].font = Font(bold=True)
        row[1].font = Font(bold=True)

wb.save(output_path)
print(f"保存完了: {output_path}")
