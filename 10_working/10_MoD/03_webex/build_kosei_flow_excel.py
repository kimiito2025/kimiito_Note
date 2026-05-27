from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

base_dir = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\03_webex")
out_path = base_dir / "4_構成概要_フロー整理_20260526_01.xlsx"

wb = Workbook()

header_fill = PatternFill("solid", fgColor="005B8F")
header_font = Font(color="FFFFFF", bold=True)
wrap = Alignment(vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def format_sheet(ws, widths):
    cols = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[cols[i - 1]].width = w

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(widths)):
        for c in row:
            c.alignment = wrap
            c.border = border

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font


# Sheet 1: スライド別フロー
ws1 = wb.active
ws1.title = "スライド別フロー"
ws1.append([
    "ページ",
    "スライドタイトル",
    "利用シナリオ",
    "主な通信",
    "送信元",
    "宛先",
    "主なポート/プロトコル",
    "CSA経由想定",
    "SSE経由フロー詳細（送信元 → 宛先 / 用途）",
    "補足",
])

slide_rows = [
    (1,  "構成概要",                                                  "全体構成",                "Webex/Entra/AWS(NTP,DNS)へのアクセス経路", "拠点内端末・LGW・Webex App", "Webex/Entra/AWS",                                          "全体図",                                                  "Mixed",  "拠点端末/Webex App → SSE → Webex/Entra/AWS\n（セキュリティポリシー適用のSaaSアクセス）",         "実線=SSE経由想定、破線=SSE経由しない想定の凡例"),
    (2,  "Webex管理ポータルへのアクセス(ログイン)",                  "管理ポータルログイン",    "Webポータルアクセス + 認証",              "保守PC/クライアントPC",      "admin.webex.com, idbroker*.webex.com, login.microsoftonline.com", "TCP/443",                                                 "Yes",    "保守PC/クライアントPC → SSE → admin.webex.com等\n（Webex管理ポータルアクセス・SAML認証の検査/制御）",         "全矢印実線。管理ポータルアクセス・SAML認証ともにSSE経由"),
    (3,  "名前解決",                                                  "DNS名前解決",             "DNS問い合わせ",                           "LGW/ハードフォン/Webex App/保守PC/構成管理サーバ", "AWS DNS",                                           "UDP/53",                                                  "Yes",    "全端末 → SSE → AWS DNS\n（DNS SecurityによるDNSクエリの検査・悪性ドメイン遮断）",                         "全端末(Agent有無問わず)の全矢印実線。SSE経由でDNS Security適用"),
    (4,  "時刻同期",                                                  "NTP時刻同期",             "NTP問い合わせ",                           "LGW/ハードフォン/Webex App/保守PC", "AWS NTP, ntp-jp.bcld.webex.com",                      "UDP/123",                                                 "Mixed",  "Webex App/PC(Agent有) → SSE → AWS NTP\n（NTP通信のSSE経由適用）",                                          "Agent有(Webex App/PC)=実線=SSE経由 / Agent無(LGW/ハードフォン)=破線=SSE非経由。MPP設定変更検討メモあり"),
    (5,  "1) ユーザ作成(Entra ID, SSE, Webex, FW)",                 "アカウント連携",          "Entra属性取得 + Webex API更新",           "構成管理サーバ",             "graph.microsoft.com, webexapis.com/v1",                   "TCP/443",                                                 "Yes",    "構成管理サーバ → SSE → graph.microsoft.com / webexapis.com\n（ユーザ同期・プロビジョニングAPIアクセスの検査/制御）", "全矢印実線。構成管理サーバからSSE経由でAPI接続"),
    (6,  "2) 認証(Entra ID, SSE, SD-WAN, FW)",                      "ソフトフォンログイン",    "Webex/Entra認証 + 設定管理",             "Webex App(Agent有/無)・LGW・ハードフォン・CUCM", "idbroker*.webex.com, login.microsoftonline.com, Webex管理IP", "TCP/443, TCP/6970, TCP/80",                   "Yes",    "全端末 → SSE → idbroker*.webex.com / login.microsoftonline.com\n（認証・MACアドレス登録・端末構成確認の検査/可視化）", "全矢印実線。Agent無デバイスのMACアドレス認証/構成確認もSSE経由"),
    (7,  "2) レジストレーション",                                     "端末登録",                "呼制御シグナリング",                      "LGW/ハードフォン/Webex App", "Webex Calling宛先レンジ",                                 "TCP/8934, TCP/5060-5080",                                 "Mixed",  "ソフトフォンPC(Agent有) → SSE → Webex Calling (TCP:8934)\n（登録シグナリングのSSE経由検査）",               "ソフトフォンPC(Agent有)TCP:8934=実線=SSE経由 / ハードフォン(Agent無)TCP:5060-5080およびSITE6=破線=SSE非経由"),
    (8,  "3) サービス確認 通話:ソフトフォン(拠点間疎通可)",          "内線疎通可通話",          "シグナリング + 通話メディア(P2P)",        "Webex App(両拠点)",          "Webex Calling",                                           "TCP/8934, UDP/8500-8699, UDP/5004,9000",                  "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。シグナリング/疎通確認/P2Pメディア全てSSE非経由"),
    (9,  "3) サービス確認 通話:ソフトフォン(拠点間疎通不可/外線通話時)", "内線疎通不可/外線",   "シグナリング + クラウド経由メディア",     "Webex App(両拠点)",          "Webex Calling",                                           "TCP/8934, UDP/8500-8699, UDP/19560-65535",                "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。外線通話はSaaS経由だがSSE非経由"),
    (10, "3) サービス確認 通話:ソフトフォン-ハードフォン(拠点間疎通可)", "ソフト/ハード混在通話", "シグナリング + 通話(P2P)",             "Webex App, ハードフォン",    "Webex Calling",                                           "TCP/8934, TCP/5060-5080, UDP/19560-19661",                "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。拠点間直接疎通あり前提だが全てSSE非経由"),
    (11, "3) サービス確認 通話:ソフトフォン-ハードフォン(拠点間疎通不可)", "ソフト/ハード混在通話", "シグナリング + クラウド中継メディア",  "Webex App, ハードフォン",    "Webex Calling",                                           "TCP/8934, TCP/5060-5080, UDP/19560-65535",                "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。拠点間直接疎通なし。全てSSE非経由"),
    (12, "3) サービス確認 通話:ハードフォン(拠点間疎通可)",           "ハードフォン間通話",      "SIPシグナリング + メディア(P2P)",         "ハードフォン(両拠点)",       "Webex Calling",                                           "TCP/5060-5080, UDP/19560-19661, UDP/5004,9000",           "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。拠点間直接疎通あり前提だが全てSSE非経由"),
    (13, "3) サービス確認 通話:ハードフォン(拠点間疎通不可/外線通話時)", "ハードフォン外線",      "SIPシグナリング + クラウド経由メディア", "ハードフォン(両拠点)",       "Webex Calling",                                           "TCP/5060-5080, UDP/19560-65535",                          "No",     "-（全通信SSE非経由）",                                                                                         "全矢印破線。外線通話はSaaS経由だがSSE非経由"),
    (14, "3) サービス確認 通話:コンタクトセンター(拠点内通話)",       "WxCC拠点内",              "Webアプリアクセス + シグナリング + 通話", "WxCC Webex App, Webex App",  "Webex Calling",                                           "TCP/443, TCP/8934, UDP/8500-8699, UDP/19560-65535",       "Mixed",  "WxCC Webex App → SSE → WxCCダッシュボード/Webex (TCP:443)\n（Webブラウザアクセスの検査/URLフィルタリング）",  "Webアプリアクセス(TCP:443)=実線=SSE経由 / シグナリング(TCP:8934)/通話メディア=破線=SSE非経由"),
    (15, "3) サービス確認 通話:コンタクトセンター(拠点間通話)",       "WxCC拠点間",              "Webアプリアクセス + シグナリング + 通話", "WxCC Webex App, Webex App",  "Webex Calling",                                           "TCP/443, TCP/8934, UDP/8500-8699, UDP/19560-65535",       "Mixed",  "WxCC Webex App → SSE → WxCCダッシュボード/Webex (TCP:443)\n（Webブラウザアクセスの検査/URLフィルタリング）",  "Webアプリアクセス(TCP:443)=実線=SSE経由 / シグナリング(TCP:8934)/通話メディア=破線=SSE非経由"),
    (16, "3) サービス確認 チャット",                                   "Webexチャット",           "チャット書き込み/表示",                  "Webex App(両拠点)",          "Webex Messaging",                                         "TCP/443",                                                 "Yes",    "Webex App(両拠点) → SSE → Webex Messaging (TCP:443)\n（チャット通信の可視化・DLP/CASB適用）",                  "全矢印実線。チャットは全通信SSE経由"),
    (17, "3) サービス確認 内部者会議",                                 "Web会議",                 "会議予約/シグナリング + 会議メディア",    "Webex App(両拠点)",          "Webex Meetings",                                          "TCP/443, UDP/9000",                                       "Mixed",  "Webex App → SSE → Webex Meetings (TCP:443)\n（会議予約・参加シグナリングの検査/URLフィルタリング）",           "会議予約/シグナリング(TCP:443)=実線=SSE経由 / 会議メディア(UDP:9000)=破線=SSE非経由"),
    (18, "3) サービス確認 通話:ソフトフォン-外線通話(オンプレ外線発信)", "PSTN発信",              "認証 + シグナリング + メディア",          "Webex App, LGW/CUCM",        "Webex Calling, 公衆網",                                   "TCP/8934, TCP/5060,5061, UDP/16384-32766, UDP/8000-48199","Mixed",  "Webex App → SSE → Webex Calling (TCP:8934)\n（通話開始前の認証シグナリングの検査）",                       "認証(TCP:8934)=実線=SSE経由 / CUCM向けシグナリング/メディア/公衆網向け通話=破線=SSE非経由"),
    (19, "3) サービス確認 通話:ソフトフォン-外線通話(災害時優先)",    "災害時PSTN",              "認証 + シグナリング + メディア",          "Webex App, Voice GW",        "Webex Calling, 公衆網",                                   "TCP/8934, TCP/5060,5061, UDP/8500-8699, UDP/19560-65535", "Mixed",  "Webex App → SSE → Webex Calling (TCP:8934)\n（通話開始前の認証シグナリングの検査）",                       "認証(TCP:8934)=実線=SSE経由 / 通話シグナリング/メディア/公衆網向け=破線=SSE非経由"),
    (20, "3) サービス確認 通話:ハードフォン-外線通話(災害時優先)",    "災害時PSTN(ハードフォン)", "SIPシグナリング + メディア",            "ハードフォン, Voice GW",     "Webex Calling, 公衆網",                                   "TCP/5060-5080, TCP/5060,5061, UDP/19560-19661, UDP/19560-65535", "Mixed", "ハードフォン/Voice GW → SSE → Webex Calling (TCP:8934)\n（通話開始前の認証シグナリングの検査）",             "認証(TCP:8934)=実線=SSE経由 / SIPシグナリング/メディア/公衆網向け=破線=SSE非経由"),
    (21, "3) サービス確認 通話:ソフトフォン-オンプレミス電話",        "オンプレPBX連携",         "認証 + シグナリング + メディア",          "Webex App, LGW/CUCM",        "Webex Calling, CUCM",                                     "TCP/8934, TCP/5060, UDP/16384-32766, UDP/19560-65535",    "Mixed",  "Webex App → SSE → Webex Calling (TCP:8934)\n（通話開始前の認証シグナリングの検査）",                       "認証(TCP:8934)=実線=SSE経由 / CUCM向けシグナリング/メディア/通話メディア=破線=SSE非経由"),
    (22, "3) サービス確認 通話:ソフトフォン-オンプレミス電話(ICE)",   "オンプレPBX連携(ICE)",    "シグナリング + ICEメディア",              "Webex App, LGW/CUCM",        "Webex Calling, CUCM",                                     "TCP/5062,8934, TCP/5060, UDP/16384-32766, UDP/8500-8699, UDP/8000-48199", "No", "-（全通信SSE非経由）",                                                                                         "全矢印破線。ICE利用時も全通信SSE非経由"),
]

for r in slide_rows:
    ws1.append(list(r))

format_sheet(ws1, [7, 36, 20, 30, 24, 38, 34, 12, 48, 28])

# Sheet 2: 通信マトリクス
ws2 = wb.create_sheet("通信マトリクス")
ws2.append([
    "通信カテゴリ",
    "代表シナリオ",
    "送信元",
    "宛先",
    "ポート/プロトコル",
    "CSA経由",
    "備考",
])

matrix = [
    ("管理ポータル",     "Webexログイン",         "保守PC/クライアントPC",     "admin.webex.com / idbroker*.webex.com / login.microsoftonline.com", "TCP/443",                                           "Yes",    "全矢印実線。ページ2"),
    ("名前解決",         "DNS",                   "LGW/端末/Webex App/構成管理サーバ", "AWS DNS",                                          "UDP/53",                                            "Yes",    "全端末(Agent有無問わず)実線。ページ3"),
    ("時刻同期",         "NTP",                   "LGW/端末/Webex App",        "AWS NTP / ntp-jp.bcld.webex.com",                  "UDP/123",                                           "Mixed",  "Agent有=実線=SSE経由 / Agent無=破線=SSE非経由。ページ4"),
    ("ユーザ同期",       "ユーザ作成",             "構成管理サーバ",             "graph.microsoft.com / webexapis.com",              "TCP/443",                                           "Yes",    "全矢印実線。ページ5"),
    ("認証",             "ソフトフォンログイン",   "Webex App/LGW/ハードフォン/CUCM", "idbroker*.webex.com / login.microsoftonline.com", "TCP/443, TCP/6970, TCP/80",                       "Yes",    "Agent無含め全矢印実線。ページ6"),
    ("レジストレーション", "端末登録",             "Webex App/ハードフォン",    "Webex Calling",                                    "TCP/8934, TCP/5060-5080",                           "Mixed",  "ソフトフォンPC(Agent有)=実線=SSE経由 / ハードフォン(Agent無)/SITE6=破線=SSE非経由。ページ7"),
    ("通話(ソフトフォン)", "拠点間可/不可",        "Webex App",                 "Webex Calling",                                    "TCP/8934, UDP/8500-8699, UDP/5004,9000, UDP/19560-65535", "No",   "全矢印破線。シグナリング/メディア全てSSE非経由。ページ8,9"),
    ("通話(ハードフォン)", "拠点間可/不可",        "ハードフォン",              "Webex Calling",                                    "TCP/5060-5080, UDP/19560-19661, UDP/19560-65535",   "No",     "全矢印破線。ページ10-13"),
    ("コンタクトセンター", "拠点内/拠点間",        "WxCC Webex App",            "Webex Calling",                                    "TCP/443, TCP/8934, UDP/8500-8699, UDP/19560-65535", "Mixed",  "Webアプリアクセス(TCP:443)=実線=SSE経由 / 通話シグナリング/メディア=破線=SSE非経由。ページ14,15"),
    ("チャット",         "メッセージング",         "Webex App",                 "Webex Messaging",                                  "TCP/443",                                           "Yes",    "全矢印実線。ページ16"),
    ("会議",             "内部者会議",             "Webex App",                 "Webex Meetings",                                   "TCP/443, UDP/9000",                                 "Mixed",  "会議シグナリング/予約(TCP:443)=実線=SSE経由 / メディア(UDP:9000)=破線=SSE非経由。ページ17"),
    ("外線通話",         "オンプレ外線/災害時優先", "Webex App/ハードフォン/LGW", "公衆網 + Webex",                                  "TCP/8934, TCP/5060,5061, UDP/16384-32766, UDP/8000-48199", "Mixed", "認証(TCP:8934)=実線=SSE経由 / 通話シグナリング/メディア/公衆網向け=破線=SSE非経由。ページ18-20"),
    ("オンプレ連携(非ICE)", "オンプレミス電話",    "Webex App/LGW/CUCM",        "CUCM + Webex",                                     "TCP/8934, TCP/5060, UDP/16384-32766, UDP/19560-65535", "Mixed", "認証(TCP:8934)=実線=SSE経由 / シグナリング/メディア=破線=SSE非経由。ページ21"),
    ("オンプレ連携(ICE)", "オンプレミス電話(ICE)", "Webex App/LGW/CUCM",        "CUCM + Webex",                                     "TCP/5062,8934, TCP/5060, UDP/16384-32766, UDP/8500-8699, UDP/8000-48199", "No", "全矢印破線。ページ22"),
]

for r in matrix:
    ws2.append(list(r))

format_sheet(ws2, [16, 22, 20, 30, 32, 10, 24])

# Sheet 3: 宛先IPレンジ
ws3 = wb.create_sheet("宛先IPレンジ")
ws3.append(["用途", "宛先/ドメイン/IP", "記載ページ"])

destinations = [
    ("管理ポータル", "admin.webex.com / user.webex.com", "2"),
    ("認証", "idbroker.webex.com / idbroker-secondary.webex.com / idbroker-b-us.webex.com / idbroker-eu.webex.com / atlas-a.wbx2.com / idbroker-ca.webex.com", "2,6"),
    ("認証", "login.microsoftonline.com", "2,6"),
    ("NTP", "ntp-jp.bcld.webex.com", "4"),
    ("API", "graph.microsoft.com / webexapis.com/v1", "5"),
    ("呼制御宛先", "23.89.0.0/16, 62.109.192.0/18, 85.119.56.0/23, 128.177.14.0/24, 128.177.36.0/24, 135.84.168.0/21, 139.177.64.0/21, 139.177.72.0/23, 144.196.0.0/16, 150.253.128.0/17, 163.129.0.0/17, 170.72.0.0/16, 170.133.128.0/18, 185.115.196.0/22, 199.19.196.0/23, 199.19.199.0/24, 199.59.64.0/21", "7-13,18-22"),
    ("チャット/会議宛先", "23.89.0.0/16, 62.109.192.0/18, 64.68.96.0/19, 66.163.32.0/19, 69.26.160.0/19, 114.29.192.0/19, 144.196.0.0/16, 150.253.128.0/18, 163.129.0.0/17, 170.133.128.0/18, 170.72.0.0/16, 173.39.224.0/19, 173.243.0.0/20, 207.182.160.0/19, 209.197.192.0/19, 210.4.192.0/20, 216.151.128.0/19", "16,17"),
]

for r in destinations:
    ws3.append(list(r))

format_sheet(ws3, [18, 105, 16])

# Sheet 4: 作成情報
ws4 = wb.create_sheet("作成情報")
ws4.append(["項目", "値"])
ws4.append(["作成日", date.today().isoformat()])
ws4.append(["基準資料", "4.構成概要.pdf (03_webex配下)"])
ws4.append(["整理対象", "全22ページのフロー概要"])
ws4.append(["凡例", "実線: SSE(CSA)経由想定 / 破線: SSE経由しない想定"])
ws4.append(["出力先", str(out_path)])

format_sheet(ws4, [22, 100])

wb.save(out_path)
print(f"保存完了: {out_path}")
