from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

webex_port_dir = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\03_webex\webex_ports_extracted\webexサービス必要ポート数")
output_path = Path(r"c:\Users\kimiito\OneDrive - Cisco\Documents\700 - github\kimiito_Note\10_working\10_MoD\03_webex\Webex_ポート使用分析_20260526_02.xlsx")

USECASES = [
    ("Baseline",       "baseline_tcp.txt",      "baseline_udp.txt",      "アプリ起動のみ（ログイン済み・待機状態）"),
    ("Calling",        "calling_tcp.txt",        "calling_udp.txt",       "Webex Calling 通話中"),
    ("Meeting All",    "meeting_all_tcp.txt",    "meeting_all_udp.txt",   "Webex Meeting（音声・カメラ・画面共有）"),
    ("Meeting Audio",  "meeting_audio_tcp.txt",  "meeting_audio_udp.txt", "Webex Meeting（音声のみ）"),
    ("Meeting A/V",    "meeting_av_tcp.txt",     "meeting_av_udp.txt",    "Webex Meeting（音声・カメラ）"),
    ("Messaging Chat", "messaging_chat_tcp.txt", "messaging_chat_udp.txt","メッセージ送受信中"),
    ("Messaging Idle", "messaging_idle_tcp.txt", "messaging_idle_udp.txt","アプリ起動・メッセージ待機"),
]

def parse_netstat(filepath):
    local_ports = set()
    remote_ports = set()
    listening_ports = set()
    with open(filepath, encoding="utf-16") as f:
        for line in f:
            m = re.search(r"UDP\s+\S+:(\d+)\s+(\S+):(\d+|\*)", line)
            if m:
                local_ports.add(int(m.group(1)))
                if m.group(3) != "*":
                    remote_ports.add(int(m.group(3)))
                continue
            m = re.search(r"TCP\s+\S+:(\d+)\s+(\S+):(\d+)\s+(\S+)", line)
            if m:
                local_ports.add(int(m.group(1)))
                if m.group(4) == "LISTENING":
                    listening_ports.add(int(m.group(1)))
                else:
                    remote_ports.add(int(m.group(3)))
    return {"local": sorted(local_ports), "remote": sorted(remote_ports), "listening": sorted(listening_ports)}

port_data = {}
for name, tf, uf, _ in USECASES:
    t = parse_netstat(webex_port_dir / tf)
    u = parse_netstat(webex_port_dir / uf)
    port_data[name] = {"tcp_local": t["local"], "tcp_remote": t["remote"], "tcp_listening": t["listening"],
                       "udp_local": u["local"], "udp_remote": u["remote"]}

OFFICIAL_SERVICES = [
    ("コールシグナリング (SIP TLS)",
     "Webex App -> Webex Calling Cloud へのSIPシグナリング\n  5062: 証明書ベーストランク\n  8934: 登録ベーストランク/App",
     {5062, 8934}, set(), ""),
    ("コールメディア (STUN/SRTP宛先5004)",
     "Webex CloudへのSRTP音声・ビデオメディア転送\n  宛先ポート: 5004 (UDP推奨)\n  クライアントのソースは動的ポートを使用",
     set(), {5004, 9000}, ""),
    ("コールメディア SRTP 音声ソース (App QoS有効時)",
     "Webex App 音声SRTPのソースポート範囲\n  ★QoS(GPO)有効化が必要。未設定環境では動的ポート(49152-65535)を使用\n  Windows: Ciscoアカウントチームへ有効化依頼が必要",
     set(), set(range(8500, 8600)),
     "QoS有効時のみ（GPO/企業設定必要）\n今回の検証環境では未使用（OSの動的ポートを使用）"),
    ("コールメディア SRTP ビデオソース (App QoS有効時)",
     "Webex App ビデオSRTPのソースポート範囲\n  ★QoS(GPO)有効化が必要。未設定環境では動的ポートを使用",
     set(), set(range(8600, 8700)),
     "QoS有効時のみ（GPO/企業設定必要）\n今回の検証環境では未使用（OSの動的ポートを使用）"),
    ("コールメディア SRTP (MPP/Roomデバイス)",
     "MPP電話/Room Series のSRTPメディアソースポート\n  Room Series: 音声52050-52099, ビデオ52200-52299",
     set(), set(range(19560, 19662)), "Webex デバイス（MPP電話など）"),
    ("Webex App ローカル LISTEN (5002/8933)",
     "Webex App がローカルでLISTENするポート\n  netstatでLISTENING状態として確認可能",
     {5002, 8933}, {8933}, ""),
    ("HTTPS / シグナリング / 認証",
     "Webex Cloud へのHTTPS通信\n  認証・設定・Control Hub・シグナリング",
     {443}, set(), ""),
    ("Webex App セキュアウェブ (8443)",
     "IDブローカー認証・Webex App構成（代替ポート）",
     {8443}, set(), ""),
    ("デバイス設定/ファームウェア管理",
     "Cisco MPP/ATAデバイスのファームウェア・設定\n  6970: ファームウェアダウンロード",
     {443, 6970, 80}, set(), "Webex デバイスのみ"),
    ("DNS 名前解決",
     "Webex Cloud サービスのFQDN/SRV解決",
     {53}, {53}, ""),
    ("NTP 時刻同期",
     "デバイス時刻同期（セキュア登録に必須）",
     set(), {123}, "Webex デバイスのみ"),
    ("CScan ネットワーク準備確認",
     "Webex Calling ネットワーク準備確認ツール\n  TCP: 443, 8934 / UDP: 19569-19760",
     {443, 8934}, set(range(19569, 19761)), ""),
    ("プッシュ通知 (APNS/FCM)",
     "モバイルデバイスへの着信・メッセージ通知",
     {443, 2197, 5223, 5228, 5229, 5230}, set(), "モバイルデバイスのみ"),
]

all_dst_tcp = set()
all_dst_udp = set()
for _, _, dt, du, _ in OFFICIAL_SERVICES:
    all_dst_tcp |= dt
    all_dst_udp |= du

wb = Workbook()
header_fill  = PatternFill("solid", fgColor="005B8F")
header_font  = Font(color="FFFFFF", bold=True, size=11)
note_fill    = PatternFill("solid", fgColor="FFF2CC")
wrap         = Alignment(vertical="top", wrap_text=True)
center       = Alignment(vertical="center", horizontal="center", wrap_text=True)
thin         = Side(style="thin", color="D9D9D9")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws):
    for c in ws[1]:
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

uc_names = [u[0] for u in USECASES]

# Sheet1
ws1 = wb.active
ws1.title = "ポート使用分析"
ws1.append(["ユースケース","説明","TCP\nローカル数","TCP\n宛先数","TCP\nLISTEN数","UDP\nローカル数","UDP\n宛先数","TCP宛先\n（Webex公式一致）","UDP宛先\n（Webex公式一致）","一致した公式サービス"])
hdr(ws1)
for name, _, _, desc in USECASES:
    d = port_data[name]
    trs, urs = set(d["tcp_remote"]), set(d["udp_remote"])
    mt = sorted(trs & all_dst_tcp)
    mu = sorted(urs & all_dst_udp)
    svcs = []
    for sn, _, dt, du, _ in OFFICIAL_SERVICES:
        th = bool(trs & dt)
        uh = bool(urs & du) if len(du) <= 20 else any(p in urs for p in list(du)[:200])
        if th or uh: svcs.append(sn)
    ws1.append([name, desc, len(d["tcp_local"]), len(d["tcp_remote"]), len(d["tcp_listening"]),
                len(d["udp_local"]), len(d["udp_remote"]),
                ", ".join(map(str, mt)) or "N/A", ", ".join(map(str, mu)) or "N/A", "\n".join(svcs)])
for col, w in zip("ABCDEFGHIJ", [18,38,11,11,11,11,11,30,20,44]):
    ws1.column_dimensions[col].width = w
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for c in row: c.alignment = wrap; c.border = border

# Sheet2
ws2 = wb.create_sheet("実宛先ポート一覧")
ws2.append(["ユースケース","プロトコル","サーバー宛先ポート（実測値）","宛先数"])
hdr(ws2)
for name, _, _, _ in USECASES:
    d = port_data[name]
    ws2.append([name, "TCP", ", ".join(map(str, d["tcp_remote"])) or "(なし)", len(d["tcp_remote"])])
    ws2.append([name, "UDP", ", ".join(map(str, d["udp_remote"])) or "(なし)", len(d["udp_remote"])])
for col, w in zip("ABCD", [18, 10, 80, 8]):
    ws2.column_dimensions[col].width = w
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for c in row: c.alignment = wrap; c.border = border

# Sheet3
ws3 = wb.create_sheet("Cisco公式ポート一覧")
h3 = ["サービス名","用途","宛先TCP\nポート","宛先UDP\nポート","注記（QoS等）"] + uc_names + ["出典"]
ws3.append(h3)
hdr(ws3)
ref = "Cisco Webex Calling ポート参照情報\nhttps://help.webex.com/ja-jp/article/b2exve"
for svc, desc, dt, du, note in OFFICIAL_SERVICES:
    ts = ", ".join(map(str, sorted(dt))) if dt else "-"
    if du:
        su = sorted(du); us = f"{su[0]}-{su[-1]}" if len(su) > 5 else ", ".join(map(str, su))
    else:
        us = "-"
    uc_cols = []
    for un in uc_names:
        d = port_data[un]
        trs, urs = set(d["tcp_remote"]), set(d["udp_remote"])
        th = bool(trs & dt)
        uh = bool(urs & du) if len(du) <= 20 else any(p in urs for p in list(du)[:200])
        uc_cols.append("✔" if (th or uh) else "")
    ws3.append([svc, desc, ts, us, note or "-"] + uc_cols + [ref])
for col, w in zip("ABCDE", [26, 46, 16, 16, 44]):
    ws3.column_dimensions[col].width = w
for i in range(len(uc_names)):
    ws3.column_dimensions[chr(ord("F")+i)].width = 14
ws3.column_dimensions[chr(ord("F")+len(uc_names))].width = 46
for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    for c in row: c.alignment = wrap; c.border = border
    if row[4].value and row[4].value != "-":
        for c in row: c.fill = note_fill

# Sheet4
ws4 = wb.create_sheet("作成情報")
info = [
    ["項目","内容"],
    ["作成日", date.today().isoformat()],
    ["基準データ", "Webex PC クライアント ネットワーク監視結果（netstat）全7ユースケース"],
    ["ポート分析の方針", "「宛先ポート（サーバー側）」を分析対象とした。クライアントのソースポートは動的（OS割り当て）"],
    ["8500-8699 未検出の理由",
     "Cisco公式では音声:8500-8599、ビデオ:8600-8699 はWebex AppのSRTPソースポート範囲として定義されているが、\n"
     "これはQoS設定（Windows GPOまたはCiscoアカウントチームによる有効化）が必要な機能。\n"
     "未設定の場合はOSが動的ポート（通常49152-65535）をソースポートとして割り当てるため、\n"
     "今回の検証環境（QoS未設定）では8500-8699は使用されず、動的ポートが使われた。\n"
     "有効化手順: https://help.webex.com/ja-jp/3btu5u/"],
    ["公式ポート出典", "Cisco Webex Calling ポート参照情報\nhttps://help.webex.com/ja-jp/article/b2exve"],
    ["対象ユースケース", ", ".join(uc_names)],
]
for r in info: ws4.append(r)
ws4.column_dimensions["A"].width = 26
ws4.column_dimensions["B"].width = 80
for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row):
    for c in row: c.alignment = wrap; c.border = border
    if row[0].row == 1:
        row[0].font = Font(bold=True); row[1].font = Font(bold=True)

wb.save(output_path)
print(f"保存完了: {output_path}")
print("\n--- 宛先ポートサマリ ---")
for name, _, _, _ in USECASES:
    d = port_data[name]
    print(f"{name:18} TCP:{sorted(set(d['tcp_remote']))}  UDP:{sorted(set(d['udp_remote']))}")
